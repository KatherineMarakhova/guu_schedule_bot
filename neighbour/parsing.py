import requests
import schedule
import time
import wget
from bs4 import BeautifulSoup
import os
import shutil
from pathlib import Path
import telebot
from openpyxl import *
from openpyxl.utils import range_boundaries
import sys
sys.path.append('../../')
import config

with open('tokens.txt', 'r') as f:
    lines = f.readlines()
    token = lines[0]
    makareshka_token = lines[1]
    print(f'marareshka: {makareshka_token}')
# for line in f:
#     line

bot = telebot.TeleBot(config.token)

# НЕЗАБЫВАТЬ ЗАПУСКАТЬ ЭТОТ ФАЙЛ ЧЕРЕЗ nohup python parsing.py

"""
Что делает этот файл:
Когда подходит запланированное время он
1. удаляет целиком папку со старыми файлами
2. создает новую папку
3. последовательно скачивает и обрабатывает новые файлы
4. отправляет сообщение разработчику, что файлы обновились
"""

#Скачивает файл относительно курса бакалавриата
def get_file(course):
    url = 'https://guu.ru/студентам/расписание-сессий/schedule/'
    response = requests.get(url)
    bs = BeautifulSoup(response.text,"lxml")
    rows = bs.find_all('a', class_ = "doc-unit odd")

    for row in rows:
        link = row.attrs["href"]
        if link == 'None': continue
        # print(link)
        if str(link).find(f'{course}-курс-бакалавриат') != -1:
            # response = requests.get(link, '../files')
            wget.download(link, '../files/')
            idx = str(link).find(f'{course}-курс-бакалавриат')
            path = link[idx:]
            return f'../files/{path}'

#Обработка файла: разделение объединенных ячеек
def unmerge_all_cells(path):
    workbook = load_workbook(path)
    n = len(workbook.sheetnames)
    for i in range(n):
        sheet = workbook.worksheets[i]
        set_merges = sorted(sheet.merged_cells.ranges.copy())
        for cell_group in set_merges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        workbook.save(path)

#Обработка файла: разделение объединенных институтов, вынос одного на другой лист
def unmerge_institutes(path):
    workbook = load_workbook(path)
    for s in workbook.sheetnames:
        if s.find(',') != -1:  # название нашего листа 'ИУПСиБК, ИИС 4 курс'
            sheet1 = workbook[s]

            t = s.find('курс') - 2  # вычленяем курс из названия листа
            z = s.find(',')  # находим позицию запятой
            fname = s[:z] + ' ' + s[t:]  # иупсибк
            sname = s[z + 2:]  # иис
            # Теперь лист на котором мы были будет называться как sname, а новый как fname
            sheet1.title = sname  # переименовыем в ИИС 4 курс
            workbook.create_sheet(fname)  # создаем лист ИУПСиБК 4 курс
            workbook.save(path)  # страхуемся, сохраняем наш док

            # переопределяем листы, так четче видно что где
            sheet1 = workbook[sname]  # иис тут заполнено
            sheet2 = workbook[fname]  # иупсибк тут пусто

            last_inst_name = ''
            y, x = get_indexes(sheet1, 'ИНСТИТУТ')
            for i in range(1, sheet1.max_column):
                # val = sheet1[y][i].value
                val = sheet1.cell(y, i).value
                if val != 'None':
                    last_inst_name = val

            inst_idx = get_indexes(sheet1, last_inst_name)  # индекс с которого начинается второй институт(ИИС)

            # print(f'Последний институт {last_inst_name}, находится {inst_idx}')

            # заполняем новый лист(ИУПСиБК)
            for i in range(1, sheet1.max_row):
                for j in range(1, inst_idx[1]):
                    sheet2.cell(i, j).value = sheet1.cell(i, j).value

            # print(f'inst_idx[1]: {inst_idx[1]}')
            # удлаляем ненужные столбцы с исходного листа
            sheet1.delete_cols(idx=5, amount=(inst_idx[1] - 5))  # тут пока костыль в виде 4 - именно столько столбцов нужно отступить слева
            # надо будет написать функцию добывающую этот индекс, чтобы было гибко
            workbook.save(path)

#Вспомогательная функция получения индекса элемента
def get_indexes(sheet, category):
    category = (category.strip()).lower()
    for i in range(1, sheet.max_row):
        for j in range(1, sheet.max_column+1):
            val = (str(sheet.cell(i, j).value).strip()).lower()
            if (val == category):
                return (i, j)

#Главная функция, активирующая всю работу
def update_docs():
    path = '../files'
    try:
        shutil.rmtree(path)
        # print("Папка удалена.")
    except OSError as error:
        print(f"Возникла ошибка: {error}")
        bot.send_message(chat_id = config.makareshka, text =f"Возникла ошибка: {error}")
    os.mkdir(path)
    # print("Папка создана.")

    with Path(r"../files") as direction:

        for i in range(1, 5):
            # filename = str(i) + "-курс-бакалавриат*.xlsx"
            path = get_file(i)  # скачиваем новый файл
            # print(f'path: {path}')
            unmerge_all_cells(path)
            unmerge_institutes(path)
        sec = time.time()
        struct = time.localtime(sec)
        t = time.strftime('%d.%m.%Y %H:%M', struct)
        bot.send_message(chat_id=config.makareshka, text ='я обновил расписание')

#Часть, отвечающая за своевременный запуск кода
schedule.every().day.at("03:00").do(update_docs)
while True:
    schedule.run_pending()
    time.sleep(1) # wait one minute

# update_docs()



