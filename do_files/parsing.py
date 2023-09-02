import requests
import time
from bs4 import BeautifulSoup
import os
import shutil
from pathlib import Path
import telebot
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import config

bot = telebot.TeleBot(config.token)

"""
Что делает этот скрипт:
Когда подходит запланированное время он
1. удаляет целиком папку со старыми файлами
2. создает новую папку
3. последовательно скачивает и обрабатывает новые файлы
4. отправляет сообщение разработчику, что файлы обновились
"""

def get_file(course):
    url = 'https://guu.ru/student/schedule/'
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'}
    response = requests.get(url, headers=headers)
    print(response)
    bs = BeautifulSoup(response.text,"lxml")
    rows = bs.find_all('a')
    print(rows)
    for row in rows:
        link = row.attrs["href"]
        if link == 'None': continue
        if str(link).find(f'{course}-курс-бакалавриат') != -1:
            os.system(f"wget {link} -P ../files")
            idx = str(link).find(f'{course}-курс-бакалавриат')
            path = link[idx:]
            return f'../files/{path}'


#Обработка файла: разделение объединенных ячеек
def unmerge_all_cells(path):
    workbook = load_workbook(path)
    n = len(workbook.sheetnames)
    for i in range(n):
        sheet = workbook.worksheets[i]
        set_merges = sorted(sheet.merged_cells.ranges.copy())
        for cell_group in set_merges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        workbook.save(path)
        workbook.close()


#Обработка файла: разделение объединенных институтов, вынос одного на другой лист
def unmerge_institutes(path):
    workbook = load_workbook(path)
    for s in workbook.sheetnames:
        if s.find(',') != -1:               # название нашего листа 'ИУПСиБК, ИИС 4 курс'
            sheet1 = workbook[s]

            t = s.find('курс') - 2          # вычленяем курс из названия листа
            z = s.find(',')                 # находим позицию запятой
            fname = s[:z] + ' ' + s[t:]     # иупсибк
            sname = s[z + 2:]               # иис
            # Теперь лист на котором мы были будет называться как sname, а новый как fname
            sheet1.title = sname            # переименовыем в ИИС 4 курс
            workbook.create_sheet(fname)    # создаем лист ИУПСиБК 4 курс
            workbook.save(path)             # сохраняем наш док

            # переопределяем листы, так четче видно что где
            sheet1 = workbook[sname]        # иис тут заполнено
            sheet2 = workbook[fname]        # иупсибк тут пусто

            last_inst_name = ''
            y, x = get_indexes(sheet1, 'ИНСТИТУТ')
            for i in range(1, sheet1.max_column):
                val = sheet1.cell(y, i).value
                if val:
                    last_inst_name = val

            inst_idx = get_indexes(sheet1, last_inst_name)

            # заполняем новый лист
            for i in range(1, sheet1.max_row):
                for j in range(1, inst_idx[1]):
                    sheet2.cell(i, j).value = sheet1.cell(i, j).value
            sheet1.delete_cols(idx=5, amount=(inst_idx[1] - 5))
            workbook.save(path)
            workbook.close()


#Вспомогательная функция получения индекса элемента
def get_indexes(sheet, category):
    category = (category.strip()).lower()
    for i in range(1, sheet.max_row):
        for j in range(1, sheet.max_column+1):
            val = (str(sheet.cell(i, j).value).strip()).lower()
            if (val == category):
                return (i, j)


#Главная функция, активирующая всю работу
def update_docs():
    path = '../files'
    try:
        shutil.rmtree(path)
    except OSError as error:
        print(f"Возникла ошибка: {error}")
        bot.send_message(chat_id = '479601165', text = f"Возникла ошибка: {error}")
    os.mkdir(path)

    with Path(r"../files") as direction:
        for i in range(1, 5):
            path = get_file(i)
            unmerge_all_cells(path)
            unmerge_institutes(path)

while True:
    sec = time.time()
    struct = time.localtime(sec)
    t = time.strftime('%H:%M', struct)
    if t == '04:00':
        update_docs()

        with open('upd_logs.txt', 'w') as logs_file:
            date = time.strftime('%d %B %H:%M')
            logs_file.write(f'Файлы с расписанием были обновлены. {date}')
            bot.send_message(chat_id='479601165', text=f'Файлы с расписанием были обновлены. {date}')
