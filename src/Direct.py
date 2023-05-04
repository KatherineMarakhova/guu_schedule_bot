from openpyxl import load_workbook
from pathlib import Path

class Direct:
    course = ''
    path = ''
    list_insts = []         # список институтов
    inst = ''               # выбранный институт
    list_edup = []
    edup = ''
    list_napr = []
    napr = ''
    list_groups = []
    group = ''

    def setget_chat_id(self, chat_id):
        """
            Получение строки и/или добавлние новой относительно chat-id
            Функция возвращает индекс строки, если нашла id, и, если не нашла, добавлет и возвращает индекс новой строки
        """
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        for i in range(1, sh.max_row + 1):
            val = sh.cell(row=i, column=1).value
            if val == chat_id:
                wb.close()
                return i

        new_row = sh.max_row + 1
        cell = sh.cell(row=new_row, column=1)
        cell.value = chat_id
        wb.save(filepath)
        self.setget_chat_id()  # рекурсия будет работать пока не добавится этот новый id

    def set_course(self, course, chat_id):
        self.course = course
        self.path = self.get_file_path(course)

        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'

        wb = load_workbook(filepath)
        sh = wb['Лист1']
        cell = sh.cell(row=user_row, column=2)
        cell.value = course
        wb.save(filepath)
        wb.close()

    def set_inst(self, inst, chat_id):
        self.inst = inst

        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        cell = sh.cell(row=user_row, column=3)
        cell.value = inst
        wb.save(filepath)
        wb.close()

    def set_napr(self, napr, chat_id):
        self.napr = napr

        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        cell = sh.cell(row=user_row, column=4)
        cell.value = napr
        wb.save(filepath)
        wb.close()

    def set_edup(self, edup, chat_id):
        self.edup = edup

        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        cell = sh.cell(row = user_row, column=5)
        cell.value = edup
        wb.save(filepath)
        wb.close()

    def set_group(self, group, chat_id):
        self.group = group

        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        cell = sh.cell(row=user_row, column=6)
        cell.value = group
        wb.save(filepath)
        wb.close()

    def get_list_inst(self, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)

        full_inst_list = wb.sheetnames
        # Иногда попадаются файлы с лишними страницами, тут их удаляем
        for s in full_inst_list:
            if s.startswith('3') or s == 'None': # если строка начинается с "3". в файле с расписанием 3 курса строки не начинаются с 3
                full_inst_list.remove(s)
        self.list_insts = full_inst_list

    # Получение списка направлений
    def get_list_napr(self, chat_id):
        # 1. определяем строку с направлениями
        # 2. определяем столбец с которого начинаются наименования
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        curr = (self.next_idx_cat(sheet, 'НАПРАВЛЕНИЕ', 'НАПРАВЛЕНИЕ', chat_id))      # строка и столбец

        temp = ''
        list_napr = []

        for j in range(curr[1], sheet.max_column+1):                           # сдвигаемся вправо на два элемента(ЭТО НАДО ИСПРАВИТЬ)
              name = str(sheet.cell(curr[0],j).value).strip()
              if (name != '' and name != temp and name!='None'):
                   list_napr.append(name)
              else:
                   continue
              temp = name
        self.list_napr = list_napr

    # Получение списка образовательных программ
    def get_list_edup(self, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        row = self.get_indexes(sheet, 'Образовательная программа')[0]            # рабочая строка
        coll = self.get_indexes_cat(sheet, data[2], 'направление', chat_id)[1]            # столбец где начинается направление
        next = self.next_idx_cat(sheet, data[2], 'направление', chat_id)[1]             # столбец где оно заканчивается
        list_edup = []
        temp = ''
        for j in range(coll, next):  # идем по столбцам
            name = str(sheet.cell(row, j).value).strip()
            if (name != '' and name != temp and name != 'None'):
                list_edup.append(name)
            else:
                continue
            temp = name  # сохраняем значение для проверки на объединенную ячейку

        self.list_edup = list_edup

    # Получение списка групп
    def get_list_group(self, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        row = (self.get_indexes(sheet, '№ учебной группы'))[0]                           # рабочая строка
        coll = (self.get_indexes_cat(sheet, data[3], 'Образовательная программа', chat_id))[1]  # столбец где начинается обр прог
        next = (self.next_idx_cat(sheet, data[3], 'Образовательная программа', chat_id))[1]     # столбец где заканчивается обр прог

        groups = []
        for j in range(coll, next):
            group = sheet.cell(row, j).value
            if group == "None": continue
            groups.append(group)
        self.list_groups = groups

    def check_name_in_list(self, name, somelist):
        for i in somelist:
            if i.find(name) != -1:
                return True
        return False

    # БЛОК ПОЛУЧЕНИЯ И ОБРАБОТКИ ФАЙЛА ==========================================================

    # Получение пути интересующего нас файла
    def get_file_path(self, course):
        with Path(r"../files") as direction:
            for f in direction.glob(str(course) + "-курс-бакалавриат*.xlsx"):
                if f: return f

    # Получение индекса(строка, столбец). Используется для значений Институт, направление, образовательная программа и тд.(категории)
    def get_indexes(self, sheet, category):
        category = (category.strip()).lower()
        for i in range(1, sheet.max_row):
            for j in range(1, sheet.max_column+1):
                val = (str(sheet.cell(i, j).value).strip()).lower()
                if (val == category):
                    return (i, j)

    # Получение индекса названия относительно категории(института, направления и тд.)
    def get_indexes_cat(self, sheet, name, category, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        name = (name.strip()).lower()
        y, x = self.get_indexes(sheet, category)    # нам нужна только строка! те у
        for i in range(y, sheet.max_row):
            for j in range(1, sheet.max_column+1):
                val = str(sheet.cell(i, j).value).strip().lower()

                if (val == name):
                    return (i, j)

    # Получение индекса другого элемента по строке(не равного по названию)
    def next_idx_cat(self, sheet, name, category, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        row = self.get_indexes(sheet, category)[0]      #строка
        name = name.strip().lower()
        coll = self.get_indexes(sheet, name)[1]              #столбец

        if coll == sheet.max_column: return (row, coll+1)

        for c in range(coll, sheet.max_column+1):  # строка не меняется
            val = str(sheet.cell(row, c).value).strip().lower()
            if val == name:
                if c == sheet.max_column:
                    return (row, sheet.max_column + 1)            #если это и есть наш последний элемент, то возращем его + 1
                continue
            elif val != name: return (row, c)
        return

    # БЛОК ВЫВОДА РАСПИСАНИЯ ======================================================================
    def get_scd_full(self, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        answer = f'Расписание для {data[3].capitalize()} {data[0]}-{data[4]}\n'
        # получаем граничные индексы
        day = self.get_indexes(sheet, 'Понедельник')
        d_idx = int(day[1])
        row = day[0]                                                                        # строка
        coll = self.get_indexes_cat(sheet, data[3], 'Образовательная программа', chat_id)[1]  # столбец
        coll = coll + int(data[4]) - 1                                                   # столбец относительно группы

        temp = ''
        lesson = ''
        for i in range(row, sheet.max_row):

            day = str(sheet.cell(i, d_idx).value)
            time = str(sheet.cell(i, d_idx+1).value)
            even = str(sheet.cell(i, d_idx+2).value)
            sbj = str(sheet.cell(i, coll).value)

            if (day.lower() == 'none'): break

            if sbj == 'None': sbj = 'Занятий нет'

            dash = '--------------------------------------️'

            if temp != day:
                spaces = ''
                n = len(dash) - len('❗️' + day + '❗️') - 1
                for i in range(n):
                    spaces += ' '
                answer += f'{dash}\n❗️{day}❗{spaces}|️\n{dash}\n'
                lesson = 1
                answer += (f'{lesson}📍{time}\n')
                temp = day
                lesson += 1

            if i % 2 != 0:
                answer += (f'{lesson}📍{time}\n- {even.capitalize()}\n {sbj}\n\n')
                lesson += 1
            else:
                answer += (f'- {even.capitalize()}\n {sbj}\n\n')

        return answer

    def get_scd_even(self, eveness, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        answer = f'Расписание для {data[3].capitalize()} {data[0]}-{data[4]}\n'
        answer += f'{eveness.capitalize()} неделя\n'

        # получаем граничные индексы
        day = self.get_indexes(sheet, 'Понедельник')
        d_idx = int(day[1])
        row = day[0]  # строка
        coll = self.get_indexes_cat(sheet, data[3], 'Образовательная программа', chat_id)[1]  # столбец
        coll = coll + int(data[4]) - 1  # столбец относительно группы
        temp = ''
        lesson = ''

        for i in range(row, sheet.max_row):
            day = str(sheet.cell(i, d_idx).value)
            if (day.lower() == 'none'): break

            time = str(sheet.cell(i, d_idx+1).value)
            even = str(sheet.cell(i, d_idx+2).value)
            sbj = str(sheet.cell(i, coll).value)

            if sbj == 'None': sbj = 'Занятий нет'
            dash = '--------------------------------------️'
            if temp != day:
                spaces = ''
                n = len(dash) - len('❗️' + day + '❗️') - 1
                for i in range(n):
                    spaces += ' '
                answer += f'{dash}\n❗️{day}❗{spaces}|️\n{dash}\n'
                lesson = 1
                temp = day
            if even.strip().lower() == eveness.strip().lower():
                answer += (f'{lesson}📍{time}\n{sbj}\n\n')
                lesson += 1
        return answer

    def get_scd_weekday(self, weekday, chat_id):
        data = self.get_userdata(chat_id)
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        sheet = wb[data[1]]

        answer = f'Расписание для {data[3]} {data[0]}-{data[4]}\n'

        spaces = ''
        dash = '--------------------------------------️'
        n = len(dash) - len('❗️' + weekday.upper() + '❗️') - 1
        for i in range(n):
            spaces += ' '
        answer += f'{dash}\n❗️{weekday.upper()}❗{spaces}|️\n{dash}\n'

        # получаем граничные индексы
        day = self.get_indexes(sheet, 'Понедельник')
        d_idx = int(day[1])
        row = day[0]  # строка
        coll = self.get_indexes_cat(sheet, data[3], 'Образовательная программа', chat_id)[1]  # столбец
        coll = coll + int(data[4]) - 1  # столбец относительно группы

        temp = ''
        lesson = 1
        for i in range(row, sheet.max_row):

            day = str(sheet.cell(i, d_idx).value)
            if (day.lower() == 'none'): break

            time = str(sheet.cell(i, d_idx+1).value)
            even = str(sheet.cell(i, d_idx+2).value)
            sbj = str(sheet.cell(i, coll).value)

            if sbj == 'None': sbj = 'Занятий нет'
            if weekday.lower() == day.lower():
                if i % 2 != 0:
                    answer += (f'{lesson}📍{time}\n- {even.capitalize()}\n {sbj}\n\n')
                    lesson += 1
                else:
                    answer += (f'- {even.capitalize()}\n {sbj}\n\n')
        return answer

    def clear_attributes(self):
        self.path = ''
        self.course = ''
        self.list_insts = []  # список институтов
        self.inst = ''  # выбранный институт
        self.list_edup = []
        self.edup = ''
        self.list_napr = []
        self.napr = ''
        self.list_groups = []
        self.group = ''


    def get_userdata(self, chat_id):
        user_row = self.setget_chat_id(chat_id)
        filepath = 'users_db.xlsx'
        wb = load_workbook(filepath)
        sh = wb['Лист1']
        course = sh.cell(user_row, 2).value
        inst = sh.cell(user_row, 3).value
        napr = sh.cell(user_row, 4).value
        edup = sh.cell(user_row, 5).value
        group = sh.cell(user_row, 6).value
        return (course, inst, napr, edup, group)

    def get_worksheet(self, data):
        path = self.get_file_path(data[0])
        wb = load_workbook(path)
        ws = wb[data[1]]
        return ws

